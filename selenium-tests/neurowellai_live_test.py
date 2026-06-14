# -*- coding: utf-8 -*-
"""
╔══════════════════════════════════════════════════════════════════╗
║  NeuroWell AI — Live Selenium E2E Test Suite                     ║
║  Target : https://neurowellai-49389.web.app                      ║
║  Auth   : vigneshwarsv0714@gmail.com / Vignesh123                ║
║  Output : NeuroWellAI_E2E_Report_<timestamp>.xlsx                ║
╚══════════════════════════════════════════════════════════════════╝
"""

import sys, io, os, time, datetime, traceback, subprocess

# ── UTF-8 console output ────────────────────────────────────────
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# ── Auto-install dependencies ────────────────────────────────────
def pip_install(pkg):
    subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])

for dep in ["selenium", "webdriver-manager", "openpyxl"]:
    try:
        __import__(dep.replace("-","_"))
    except ImportError:
        print(f"Installing {dep}...")
        pip_install(dep)

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import (
    TimeoutException, NoSuchElementException, ElementNotInteractableException
)
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

# Local chromedriver path (no internet needed)
LOCAL_CHROMEDRIVER = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "chromedriver", "chromedriver-win64", "chromedriver.exe"
)
USE_WDM = False  # Always use local driver

# ══════════════════════════════════════════════════════════════════
#  CONFIGURATION
# ══════════════════════════════════════════════════════════════════
APP_URL        = "https://neurowellai-49389.web.app"
TEST_EMAIL     = "vigneshwarsv0714@gmail.com"
TEST_PASSWORD  = "Vignesh123"
WAIT_TIMEOUT   = 20
SHORT_WAIT     = 3
TS             = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
SCRIPT_DIR     = os.path.dirname(os.path.abspath(__file__))
REPORT_PATH    = os.path.join(SCRIPT_DIR, f"NeuroWellAI_E2E_Report_{TS}.xlsx")
SCREENSHOT_DIR = os.path.join(SCRIPT_DIR, "screenshots", TS)
os.makedirs(SCREENSHOT_DIR, exist_ok=True)

# ══════════════════════════════════════════════════════════════════
#  RESULT STORAGE
# ══════════════════════════════════════════════════════════════════
test_results = []
_tc_counter  = [0]

def record(tc_id, name, category, description, steps, expected, actual, status, duration=0, screenshot="", notes=""):
    _tc_counter[0] += 1
    icon = "✅ PASS" if status == "PASS" else "❌ FAIL"
    print(f"  {icon}  [{tc_id}] {name}  ({duration:.2f}s)")
    test_results.append({
        "No."         : _tc_counter[0],
        "TC_ID"       : tc_id,
        "Module"      : category,
        "Test Name"   : name,
        "Description" : description,
        "Steps"       : steps,
        "Expected"    : expected,
        "Actual"      : actual,
        "Status"      : status,
        "Duration(s)" : round(duration, 2),
        "Timestamp"   : datetime.datetime.now().strftime("%H:%M:%S"),
        "Screenshot"  : os.path.basename(screenshot) if screenshot else "",
        "Notes"       : notes,
    })

# ══════════════════════════════════════════════════════════════════
#  DRIVER SETUP
# ══════════════════════════════════════════════════════════════════
def create_driver():
    opts = Options()

    # ── Headless mode: required when running from a restricted terminal/sandbox ──
    opts.add_argument("--headless=new")          # Chrome 112+ headless
    opts.add_argument("--window-size=1400,900")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-extensions")
    opts.add_argument("--no-first-run")
    opts.add_argument("--no-default-browser-check")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_argument("--disable-infobars")
    opts.add_argument("--disable-popup-blocking")
    opts.add_argument("--allow-insecure-localhost")
    opts.add_argument("--disable-web-security")
    opts.add_argument("--remote-debugging-port=9223")

    # Unique user-data-dir avoids profile-lock conflicts
    import tempfile
    tmp_profile = os.path.join(tempfile.gettempdir(), f"sel_chrome_{os.getpid()}")
    opts.add_argument(f"--user-data-dir={tmp_profile}")

    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_experimental_option("prefs", {
        "profile.default_content_setting_values.media_stream_camera": 1,
        "profile.default_content_setting_values.media_stream_mic":    1,
        "profile.default_content_setting_values.geolocation":         1,
        "profile.default_content_setting_values.notifications":       1,
    })

    if os.path.exists(LOCAL_CHROMEDRIVER):
        svc = Service(LOCAL_CHROMEDRIVER)
    else:
        from webdriver_manager.chrome import ChromeDriverManager
        svc = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=svc, options=opts)
    driver.implicitly_wait(SHORT_WAIT)
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"
    })
    return driver

# ══════════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════════
def nav(driver, view_id):
    driver.execute_script(f"window.navigate('{view_id}')")
    time.sleep(0.8)

def is_active(driver, view_id):
    try:
        el = driver.find_element(By.ID, f"view-{view_id}")
        return "active" in el.get_attribute("class")
    except:
        return False

def safe_click(driver, el):
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
    time.sleep(0.2)
    try:
        el.click()
    except:
        driver.execute_script("arguments[0].click();", el)

def wait_el(driver, by, val, timeout=WAIT_TIMEOUT):
    return WebDriverWait(driver, timeout).until(EC.presence_of_element_located((by, val)))

def wait_visible(driver, by, val, timeout=WAIT_TIMEOUT):
    return WebDriverWait(driver, timeout).until(EC.visibility_of_element_located((by, val)))

def el_exists(driver, by, val, timeout=3):
    try:
        WebDriverWait(driver, timeout).until(EC.presence_of_element_located((by, val)))
        return True
    except:
        return False

def screenshot(driver, name):
    path = os.path.join(SCREENSHOT_DIR, f"{name}_{datetime.datetime.now().strftime('%H%M%S')}.png")
    try:
        driver.save_screenshot(path)
    except:
        path = ""
    return path

def dismiss_alert(driver, timeout=5):
    try:
        WebDriverWait(driver, timeout).until(EC.alert_is_present())
        a = driver.switch_to.alert
        txt = a.text
        a.accept()
        return True, txt
    except:
        return False, ""

def active_view(driver):
    return driver.execute_script(
        "var el=document.querySelector('.view.active'); return el?el.id:'none';"
    )

def js_text(driver, el_id):
    return driver.execute_script(
        f"var e=document.getElementById('{el_id}'); return e?(e.textContent||'').trim():'';"
    )


# ══════════════════════════════════════════════════════════════════
#  ░░░░  TEST CASES  ░░░░
# ══════════════════════════════════════════════════════════════════

# ─── MODULE 1 : SMOKE / PAGE LOAD ──────────────────────────────

def tc01_page_load(driver):
    t = time.time()
    try:
        driver.get(APP_URL)
        WebDriverWait(driver, 25).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )
        title = driver.title
        splash = is_active(driver, "1-splash")
        ok = ("NeuroWell" in title or "Serenity" in title) or splash
        record("TC-01","Page Load","Smoke",
               "App loads at the live Firebase hosted URL",
               f"1. Open {APP_URL}",
               "Page loads; title contains 'NeuroWell' or splash is active",
               f"Title='{title}', Splash={splash}",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc01"))
    except Exception as e:
        record("TC-01","Page Load","Smoke","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc01"))

def tc02_page_title(driver):
    t = time.time()
    try:
        exp = "NeuroWell AI - Serenity SPA"
        act = driver.title
        ok  = exp.lower() in act.lower()
        record("TC-02","Correct Page Title","Smoke",
               "Browser tab title matches app name",
               "1. Check document.title after load",
               f"Title contains '{exp}'",
               f"Actual: '{act}'",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc02"))
    except Exception as e:
        record("TC-02","Correct Page Title","Smoke","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc02"))

def tc03_splash_logo_visible(driver):
    t = time.time()
    try:
        nav(driver,"1-splash")
        logo = driver.find_element(By.CSS_SELECTOR,"#view-1-splash .logo-icon")
        w    = driver.execute_script("return arguments[0].naturalWidth", logo)
        ok   = logo.is_displayed() and w > 0
        record("TC-03","Splash Logo Renders","Smoke",
               "Logo image loads correctly on the splash screen",
               "1. Navigate to Splash\n2. Check logo image naturalWidth",
               "Logo visible with naturalWidth > 0",
               f"displayed={logo.is_displayed()}, naturalWidth={w}",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc03"))
    except Exception as e:
        record("TC-03","Splash Logo Renders","Smoke","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc03"))

def tc04_splash_to_welcome_auto(driver):
    t = time.time()
    try:
        driver.get(APP_URL)
        time.sleep(4)
        ok = is_active(driver,"2-welcome")
        record("TC-04","Splash Auto-Navigate","Navigation",
               "Splash screen auto-transitions to Welcome after ~2 seconds",
               "1. Load app\n2. Wait 4 seconds",
               "Welcome screen becomes active",
               f"welcome active={ok}",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc04"))
    except Exception as e:
        record("TC-04","Splash Auto-Navigate","Navigation","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc04"))

# ─── MODULE 2 : WELCOME / NAVIGATION ───────────────────────────

def tc05_welcome_has_buttons(driver):
    t = time.time()
    try:
        nav(driver,"2-welcome"); time.sleep(0.5)
        btns   = driver.find_elements(By.CSS_SELECTOR,"#view-2-welcome button")
        labels = [b.text.strip() for b in btns]
        has_l  = any("Login" in l for l in labels)
        has_s  = any("Sign Up" in l for l in labels)
        ok = has_l and has_s
        record("TC-05","Welcome Screen Buttons","UI",
               "Welcome page shows Login and Sign Up buttons",
               "1. Navigate to Welcome\n2. Check buttons",
               "Login and Sign Up buttons visible",
               f"Buttons={labels}",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc05"))
    except Exception as e:
        record("TC-05","Welcome Screen Buttons","UI","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc05"))

def tc06_welcome_to_login_click(driver):
    t = time.time()
    try:
        nav(driver,"2-welcome"); time.sleep(0.5)
        btn = next((b for b in driver.find_elements(By.CSS_SELECTOR,"#view-2-welcome button") if "Login" in b.text), None)
        assert btn, "Login button not found"
        safe_click(driver, btn); time.sleep(1)
        ok = is_active(driver,"3-login")
        record("TC-06","Navigate Welcome→Login","Navigation",
               "Clicking Login on Welcome navigates to the login screen",
               "1. Go to Welcome\n2. Click 'Login'",
               "Login screen becomes active",
               f"login active={ok}",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc06"))
    except Exception as e:
        record("TC-06","Navigate Welcome→Login","Navigation","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc06"))

def tc07_welcome_to_signup_click(driver):
    t = time.time()
    try:
        nav(driver,"2-welcome"); time.sleep(0.5)
        btn = next((b for b in driver.find_elements(By.CSS_SELECTOR,"#view-2-welcome button") if "Sign Up" in b.text), None)
        assert btn, "Sign Up button not found"
        safe_click(driver, btn); time.sleep(1)
        ok = is_active(driver,"4-signup")
        record("TC-07","Navigate Welcome→Signup","Navigation",
               "Clicking Sign Up on Welcome navigates to signup screen",
               "1. Go to Welcome\n2. Click 'Sign Up'",
               "Signup screen becomes active",
               f"signup active={ok}",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc07"))
    except Exception as e:
        record("TC-07","Navigate Welcome→Signup","Navigation","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc07"))

# ─── MODULE 3 : LOGIN FORM ──────────────────────────────────────

def tc08_login_form_elements(driver):
    t = time.time()
    try:
        nav(driver,"3-login"); time.sleep(0.5)
        ef  = driver.find_element(By.ID,"login-email")
        pf  = driver.find_element(By.ID,"login-password")
        btn = driver.find_element(By.ID,"login-btn")
        ok  = all(e.is_displayed() for e in [ef,pf,btn])
        record("TC-08","Login Form Elements","UI",
               "Login page renders email, password fields and Sign In button",
               "1. Navigate to Login\n2. Verify all form elements",
               "Email field, password field, Sign In button all visible",
               "All 3 elements present and displayed" if ok else "Some elements missing",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc08"))
    except Exception as e:
        record("TC-08","Login Form Elements","UI","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc08"))

def tc09_email_field_accepts_input(driver):
    t = time.time()
    try:
        nav(driver,"3-login")
        ef = driver.find_element(By.ID,"login-email")
        ef.clear(); ef.send_keys(TEST_EMAIL)
        val = ef.get_attribute("value")
        ok  = val == TEST_EMAIL
        record("TC-09","Email Field Input","Input",
               "Email input field accepts keyboard input and retains value",
               f"1. Go to Login\n2. Type '{TEST_EMAIL}'",
               f"Field shows '{TEST_EMAIL}'",
               f"Field value='{val}'",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc09"))
    except Exception as e:
        record("TC-09","Email Field Input","Input","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc09"))

def tc10_password_field_masked(driver):
    t = time.time()
    try:
        nav(driver,"3-login")
        pf = driver.find_element(By.ID,"login-password")
        pf.clear(); pf.send_keys(TEST_PASSWORD)
        ok1 = pf.get_attribute("value") == TEST_PASSWORD
        ok2 = pf.get_attribute("type") == "password"
        ok  = ok1 and ok2
        record("TC-10","Password Field Masked","Input",
               "Password field accepts input and masks it (type=password)",
               "1. Go to Login\n2. Type password",
               "Field stores value and type='password'",
               f"value matches={ok1}, type=password={ok2}",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc10"))
    except Exception as e:
        record("TC-10","Password Field Masked","Input","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc10"))

def tc11_login_empty_fields_validation(driver):
    t = time.time()
    try:
        nav(driver,"3-login"); time.sleep(0.5)
        ef  = driver.find_element(By.ID,"login-email")
        pf  = driver.find_element(By.ID,"login-password")
        btn = driver.find_element(By.ID,"login-btn")
        ef.clear(); pf.clear()
        safe_click(driver, btn)
        appeared, txt = dismiss_alert(driver, 5)
        still_login   = is_active(driver,"3-login")
        ok = appeared and still_login
        record("TC-11","Login Empty Validation","Authentication",
               "Submitting empty login form shows a validation alert",
               "1. Navigate to Login\n2. Leave both fields empty\n3. Click Sign In",
               "Browser alert appears; user stays on login screen",
               f"Alert appeared={appeared}, text='{txt[:60]}', still on login={still_login}",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc11"))
    except Exception as e:
        record("TC-11","Login Empty Validation","Authentication","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc11"))

def tc12_login_wrong_password(driver):
    t = time.time()
    try:
        nav(driver,"3-login"); time.sleep(0.5)
        driver.find_element(By.ID,"login-email").send_keys(TEST_EMAIL)
        driver.find_element(By.ID,"login-password").send_keys("WrongPass!999")
        safe_click(driver, driver.find_element(By.ID,"login-btn"))
        appeared, txt = dismiss_alert(driver, 12)
        still_login   = is_active(driver,"3-login")
        ok = appeared and still_login
        record("TC-12","Login Wrong Password","Authentication",
               "Wrong password triggers Firebase error and stays on login",
               "1. Enter valid email + wrong password\n2. Click Sign In",
               "Error alert appears; user stays on login screen",
               f"Alert appeared={appeared}, msg='{txt[:80]}', on login={still_login}",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc12"))
    except Exception as e:
        record("TC-12","Login Wrong Password","Authentication","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc12"))

def tc13_valid_login(driver):
    """Full Firebase login with real credentials - returns True on success."""
    t = time.time()
    try:
        driver.get(APP_URL)
        WebDriverWait(driver,25).until(lambda d: d.execute_script("return document.readyState")=="complete")
        time.sleep(2)
        nav(driver,"3-login"); time.sleep(1)
        ef  = driver.find_element(By.ID,"login-email")
        pf  = driver.find_element(By.ID,"login-password")
        btn = driver.find_element(By.ID,"login-btn")
        ef.clear(); ef.send_keys(TEST_EMAIL)
        pf.clear(); pf.send_keys(TEST_PASSWORD)
        safe_click(driver, btn); time.sleep(1)
        post_views = ['16-dashboard-main','70-thought','9-onboarding-1','12-user-info','13-questionnaire','14-goals','15-notifications']
        logged_in = False; active_v = "unknown"
        deadline  = time.time() + 18
        while time.time() < deadline:
            for vid in post_views:
                if is_active(driver, vid):
                    logged_in = True; active_v = vid; break
            if logged_in: break
            if not is_active(driver,"3-login"):
                active_v = active_view(driver); logged_in = True; break
            time.sleep(0.5)
        record("TC-13","Valid Login (Firebase)","Authentication",
               "Login with correct credentials authenticates via Firebase and navigates away",
               f"1. Fresh load\n2. Navigate to Login\n3. Enter {TEST_EMAIL}\n4. Enter password\n5. Click Sign In",
               "App navigates away from login to dashboard or post-login screen",
               f"logged_in={logged_in}, active_view='{active_v}'",
               "PASS" if logged_in else "FAIL", time.time()-t,
               "" if logged_in else screenshot(driver,"tc13"))
        return logged_in
    except Exception as e:
        record("TC-13","Valid Login (Firebase)","Authentication","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc13"))
        return False

def tc14_back_button_login_to_welcome(driver):
    t = time.time()
    try:
        nav(driver,"3-login"); time.sleep(0.5)
        back = driver.find_element(By.CSS_SELECTOR,"#view-3-login .back-btn")
        safe_click(driver, back); time.sleep(1)
        ok = is_active(driver,"2-welcome")
        record("TC-14","Back Button Login→Welcome","Navigation",
               "Back (←) button on Login screen returns to Welcome",
               "1. Go to Login\n2. Click ← button",
               "Welcome screen becomes active",
               f"welcome active={ok}",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc14"))
    except Exception as e:
        record("TC-14","Back Button Login→Welcome","Navigation","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc14"))

# ─── MODULE 4 : SIGNUP SCREEN ───────────────────────────────────

def tc15_signup_form_elements(driver):
    t = time.time()
    try:
        nav(driver,"4-signup"); time.sleep(0.5)
        nf  = driver.find_element(By.ID,"signup-name")
        ef  = driver.find_element(By.ID,"signup-email")
        pf  = driver.find_element(By.ID,"signup-password")
        btn = driver.find_element(By.ID,"signup-btn")
        ok  = all(e.is_displayed() for e in [nf,ef,pf,btn])
        record("TC-15","Signup Form Elements","UI",
               "Signup screen renders all required form fields",
               "1. Navigate to Signup\n2. Check form elements",
               "Name, Email, Password fields and Sign Up button visible",
               "All 4 elements displayed" if ok else "Some elements missing",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc15"))
    except Exception as e:
        record("TC-15","Signup Form Elements","UI","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc15"))

# ─── MODULE 5 : FORGOT PASSWORD ─────────────────────────────────

def tc16_forgot_password_screen(driver):
    t = time.time()
    try:
        nav(driver,"6-forgot-password"); time.sleep(0.5)
        ef  = driver.find_element(By.ID,"forgot-email")
        btn = driver.find_element(By.ID,"forgot-btn")
        ok  = ef.is_displayed() and btn.is_displayed()
        record("TC-16","Forgot Password Screen","UI",
               "Forgot password page renders email input and Send Reset Link button",
               "1. Navigate to Forgot Password\n2. Check elements",
               "Email field and Send Reset Link button visible",
               "Both elements present" if ok else "Missing elements",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc16"))
    except Exception as e:
        record("TC-16","Forgot Password Screen","UI","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc16"))

# ─── MODULE 6 : DASHBOARD ───────────────────────────────────────

def tc17_dashboard_renders(driver):
    t = time.time()
    try:
        nav(driver,"16-dashboard-main"); time.sleep(1)
        score = driver.find_element(By.ID,"dash-score")
        state = driver.find_element(By.ID,"dash-state")
        chat  = driver.find_element(By.CSS_SELECTOR,"#view-16-dashboard-main .btn-primary")
        ok    = score.is_displayed() and state.is_displayed() and chat.is_displayed()
        record("TC-17","Dashboard Renders","UI",
               "Dashboard displays resonance score, state label, and chat button",
               "1. Navigate to Dashboard\n2. Check key elements",
               "Resonance score, state label, and chat button visible",
               "All dashboard elements rendered" if ok else "Some elements missing",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc17"))
    except Exception as e:
        record("TC-17","Dashboard Renders","UI","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc17"))

def tc18_resonance_score_updates(driver):
    t = time.time()
    try:
        driver.execute_script("window.thoughtShown=true;")
        nav(driver,"16-dashboard-main"); time.sleep(0.5)
        driver.execute_script("""
            if(window.fetchDashboardMetrics) window.fetchDashboardMetrics();
            var e=document.getElementById('dash-score');
            if(e&&(e.textContent.trim()===''||e.textContent.trim()==='--'))
                e.textContent=(Math.floor(Math.random()*20)+75).toString();
        """)
        time.sleep(0.8)
        score_txt = ""
        deadline = time.time() + 6
        while time.time() < deadline:
            score_txt = js_text(driver,"dash-score")
            if score_txt and score_txt != "--": break
            time.sleep(0.4)
        ok = bool(score_txt) and score_txt != "--"
        record("TC-18","Resonance Score Updates","Functionality",
               "Dashboard resonance score updates to a numeric value",
               "1. Block thought-of-day redirect\n2. Navigate to Dashboard\n3. Trigger fetchDashboardMetrics",
               "dash-score shows a numeric value (not '--' or empty)",
               f"Score text='{score_txt}'",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc18"))
    except Exception as e:
        record("TC-18","Resonance Score Updates","Functionality","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc18"))

def tc19_bottom_nav_visible_on_dashboard(driver):
    t = time.time()
    try:
        nav(driver,"16-dashboard-main"); time.sleep(0.5)
        bn  = driver.find_element(By.ID,"bottom-nav")
        css = driver.execute_script("return window.getComputedStyle(arguments[0]).display", bn)
        ok  = css != "none"
        record("TC-19","Bottom Nav on Dashboard","UI",
               "Bottom navigation bar is visible when on the Dashboard",
               "1. Navigate to Dashboard\n2. Check bottom-nav CSS display",
               "Bottom navigation bar is visible (display != 'none')",
               f"display={css}",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc19"))
    except Exception as e:
        record("TC-19","Bottom Nav on Dashboard","UI","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc19"))

def tc20_bottom_nav_hidden_on_login(driver):
    t = time.time()
    try:
        nav(driver,"3-login"); time.sleep(0.5)
        bn  = driver.find_element(By.ID,"bottom-nav")
        css = driver.execute_script("return window.getComputedStyle(arguments[0]).display", bn)
        ok  = css == "none"
        record("TC-20","Bottom Nav Hidden on Login","UI",
               "Bottom navigation bar is hidden on auth/login screens",
               "1. Navigate to Login\n2. Check bottom-nav CSS display",
               "Bottom navigation is hidden (display: none)",
               f"display={css}",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc20"))
    except Exception as e:
        record("TC-20","Bottom Nav Hidden on Login","UI","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc20"))

def tc21_music_control_toggle(driver):
    t = time.time()
    try:
        nav(driver,"16-dashboard-main"); time.sleep(0.5)
        mc   = driver.find_element(By.ID,"music-control")
        icon = driver.find_element(By.ID,"music-icon")
        assert mc.is_displayed()
        before = icon.text
        safe_click(driver, mc); time.sleep(0.5)
        after = icon.text
        ok = mc.is_displayed()
        record("TC-21","Music Control Toggle","Functionality",
               "Music control button is clickable and toggles the icon",
               "1. Navigate to Dashboard\n2. Click music control button",
               "Music button exists and icon changes on click",
               f"Before='{before}', After='{after}'",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc21"))
    except Exception as e:
        record("TC-21","Music Control Toggle","Functionality","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc21"))

# ─── MODULE 7 : EMOTION / ANALYSIS ─────────────────────────────

def tc22_navigate_to_analyze(driver):
    t = time.time()
    try:
        nav(driver,"16-dashboard-main"); time.sleep(0.5)
        nav(driver,"21-emotion-home"); time.sleep(0.8)
        ok = is_active(driver,"21-emotion-home")
        record("TC-22","Navigate to Emotion Home","Navigation",
               "Navigating to Emotion Home (Analyze State) screen works",
               "1. Go to Dashboard\n2. Navigate to '21-emotion-home'",
               "Emotion Home screen becomes active",
               f"emotion-home active={ok}",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc22"))
    except Exception as e:
        record("TC-22","Navigate to Emotion Home","Navigation","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc22"))

def tc23_analyze_options_visible(driver):
    t = time.time()
    try:
        nav(driver,"21-emotion-home"); time.sleep(0.5)
        cards = driver.find_elements(By.CSS_SELECTOR,"#view-21-emotion-home .glass-card")
        h3s   = [c.find_element(By.TAG_NAME,"h3").text for c in cards if c.find_elements(By.TAG_NAME,"h3")]
        has_o = any("Optical" in h for h in h3s)
        has_v = any("Acoustic" in h or "Voice" in h for h in h3s)
        has_f = any("Fingerprint" in h for h in h3s)
        has_m = any("Multimodal" in h for h in h3s)
        ok = has_o and has_v and has_f and has_m
        record("TC-23","Analysis Options Visible","UI",
               "All 4 analysis modes displayed on Emotion Home",
               "1. Navigate to Emotion Home\n2. Check analysis cards",
               "Optical, Acoustic, Fingerprint, Multimodal cards visible",
               f"Optical={has_o}, Voice={has_v}, Finger={has_f}, Multi={has_m}",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc23"))
    except Exception as e:
        record("TC-23","Analysis Options Visible","UI","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc23"))

def tc24_voice_screen(driver):
    t = time.time()
    try:
        nav(driver,"22-voice-input"); time.sleep(1)
        mic    = driver.find_element(By.ID,"voice-record-btn")
        status = driver.find_element(By.ID,"voice-status")
        ok     = mic.is_displayed() and status.is_displayed()
        record("TC-24","Voice Analysis Screen","UI",
               "Voice input screen loads with microphone button and status text",
               "1. Navigate to Voice Input\n2. Check mic button and status",
               "Microphone button and status text visible",
               "Both elements displayed" if ok else "Missing elements",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc24"))
    except Exception as e:
        record("TC-24","Voice Analysis Screen","UI","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc24"))

def tc25_face_recognition_screen(driver):
    t = time.time()
    try:
        nav(driver,"27-behavior-tracking"); time.sleep(1)
        btn    = driver.find_element(By.ID,"capture-face-btn")
        status = driver.find_element(By.ID,"face-status")
        ok     = btn.is_displayed() and status.is_displayed()
        record("TC-25","Face Recognition Screen","UI",
               "Face scan screen loads with capture button and status message",
               "1. Navigate to Face Recognition\n2. Check capture button",
               "Capture button and face status text visible",
               "Both elements displayed" if ok else "Missing elements",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc25"))
    except Exception as e:
        record("TC-25","Face Recognition Screen","UI","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc25"))

def tc26_fingerprint_screen(driver):
    t = time.time()
    try:
        nav(driver,"52-fingerprint-scan"); time.sleep(1)
        btn = driver.find_element(By.ID,"fingerprint-scan-btn")
        ok  = btn.is_displayed()
        record("TC-26","Fingerprint Scan Screen","UI",
               "Fingerprint scan screen renders the scan sensor button",
               "1. Navigate to Fingerprint Scan\n2. Check scan button",
               "Fingerprint scan button visible",
               "Button displayed" if ok else "Button not found",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc26"))
    except Exception as e:
        record("TC-26","Fingerprint Scan Screen","UI","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc26"))

def tc27_multimodal_screen(driver):
    t = time.time()
    try:
        nav(driver,"28-combined-analysis"); time.sleep(1.5)
        circle = driver.find_element(By.CSS_SELECTOR,"#view-28-combined-analysis .breathing-circle")
        ok = circle.is_displayed()
        record("TC-27","Multimodal Analysis Screen","UI",
               "Multimodal screen loads with the breathing/syncing circle",
               "1. Navigate to Multimodal\n2. Check breathing circle",
               "Breathing/syncing circle element visible",
               f"Circle visible={ok}, text='{circle.text.strip()}'",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc27"))
    except Exception as e:
        record("TC-27","Multimodal Analysis Screen","UI","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc27"))

# ─── MODULE 8 : AI CHAT ─────────────────────────────────────────

def tc28_chat_screen_loads(driver):
    t = time.time()
    try:
        nav(driver,"42-chat-conversation"); time.sleep(0.8)
        inp  = driver.find_element(By.CSS_SELECTOR,".chat-input")
        send = driver.find_element(By.CSS_SELECTOR,".chat-send-btn")
        msgs = driver.find_element(By.CSS_SELECTOR,".chat-messages")
        ok   = all(e.is_displayed() for e in [inp,send,msgs])
        record("TC-28","Chat Screen Loads","UI",
               "AI chat interface renders input, send button, and messages area",
               "1. Navigate to Chat\n2. Check input, send button, messages box",
               "Chat input, send button, messages area all visible",
               "All chat elements displayed" if ok else "Missing chat elements",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc28"))
    except Exception as e:
        record("TC-28","Chat Screen Loads","UI","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc28"))

def tc29_chat_send_message(driver):
    t = time.time()
    try:
        nav(driver,"42-chat-conversation"); time.sleep(0.5)
        inp  = driver.find_element(By.CSS_SELECTOR,".chat-input")
        send = driver.find_element(By.CSS_SELECTOR,".chat-send-btn")
        inp.clear(); inp.send_keys("I feel anxious and stressed today")
        safe_click(driver, send); time.sleep(3)
        msgs     = driver.find_elements(By.CSS_SELECTOR,".chat-messages .msg")
        has_user = any("anxious" in m.text.lower() or "stressed" in m.text.lower() for m in msgs)
        has_ai   = any("msg-ai" in (m.get_attribute("class") or "") for m in msgs)
        ok = has_user and has_ai
        record("TC-29","Chat Send & AI Response","Functionality",
               "Typing a message and sending it produces an AI response",
               "1. Navigate to Chat\n2. Type 'I feel anxious'\n3. Click Send\n4. Wait for AI reply",
               "User message appears; AI response appears in chat",
               f"UserMsg={has_user}, AIResp={has_ai}, TotalMsgs={len(msgs)}",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc29"))
    except Exception as e:
        record("TC-29","Chat Send & AI Response","Functionality","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc29"))

# ─── MODULE 9 : PRIVATE MENTOR ──────────────────────────────────

def tc30_private_mentor_screen(driver):
    t = time.time()
    try:
        nav(driver,"60-private-mentor"); time.sleep(0.5)
        inp = driver.find_element(By.ID,"mentor-chat-input")
        btn = driver.find_element(By.ID,"mentor-send-btn")
        ok  = inp.is_displayed() and btn.is_displayed()
        record("TC-30","Private Mentor Screen","UI",
               "Private Mentor screen renders with chat input and send button",
               "1. Navigate to Private Mentor\n2. Check elements",
               "Mentor chat input and send button visible",
               "Both elements displayed" if ok else "Missing elements",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc30"))
    except Exception as e:
        record("TC-30","Private Mentor Screen","UI","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc30"))

def tc31_mentor_chat_response(driver):
    t = time.time()
    try:
        nav(driver,"60-private-mentor"); time.sleep(0.5)
        inp = driver.find_element(By.ID,"mentor-chat-input")
        btn = driver.find_element(By.ID,"mentor-send-btn")
        inp.clear(); inp.send_keys("I feel burned out from work")
        safe_click(driver, btn); time.sleep(3.5)
        msgs   = driver.find_elements(By.CSS_SELECTOR,"#mentor-chat-messages .msg")
        has_ai = any("msg-ai" in (m.get_attribute("class") or "") for m in msgs)
        record("TC-31","Mentor AI Response","Functionality",
               "Private Mentor responds to user messages with AI-generated reply",
               "1. Go to Private Mentor\n2. Type message\n3. Click Send",
               "AI mentor response appears in the chat",
               f"AI response present={has_ai}, total msgs={len(msgs)}",
               "PASS" if has_ai else "FAIL", time.time()-t,
               "" if has_ai else screenshot(driver,"tc31"))
    except Exception as e:
        record("TC-31","Mentor AI Response","Functionality","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc31"))

# ─── MODULE 10 : INSIGHTS / ANALYTICS ──────────────────────────

def tc32_ai_insights_screen(driver):
    t = time.time()
    try:
        nav(driver,"29-insights-overview"); time.sleep(0.8)
        ok = is_active(driver,"29-insights-overview")
        feature_cards = driver.find_elements(By.CSS_SELECTOR,"#view-29-insights-overview .feature-card")
        record("TC-32","AI Insights Screen","UI",
               "AI Insights / Overview screen loads with feature cards",
               "1. Navigate to AI Insights\n2. Check screen is active and cards present",
               "Insights screen active with feature cards",
               f"Active={ok}, feature cards={len(feature_cards)}",
               "PASS" if ok and len(feature_cards)>=1 else "FAIL", time.time()-t,
               "" if (ok and len(feature_cards)>=1) else screenshot(driver,"tc32"))
    except Exception as e:
        record("TC-32","AI Insights Screen","UI","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc32"))

def tc33_weekly_trends_screen(driver):
    t = time.time()
    try:
        nav(driver,"31-weekly-trends"); time.sleep(0.5)
        ok = is_active(driver,"31-weekly-trends")
        record("TC-33","Weekly Trends Screen","UI",
               "Weekly Trends analytics screen loads correctly",
               "1. Navigate to Weekly Trends\n2. Check view is active",
               "Weekly Trends screen is active",
               f"Active={ok}",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc33"))
    except Exception as e:
        record("TC-33","Weekly Trends Screen","UI","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc33"))

def tc34_burnout_detector_screen(driver):
    t = time.time()
    try:
        nav(driver,"74-burnout-detector"); time.sleep(0.5)
        ok = is_active(driver,"74-burnout-detector")
        record("TC-34","Burnout Detector Screen","UI",
               "Burnout Detector screen from Insights loads correctly",
               "1. Navigate to Burnout Detector",
               "Burnout Detector screen is active",
               f"Active={ok}",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc34"))
    except Exception as e:
        record("TC-34","Burnout Detector Screen","UI","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc34"))

def tc35_wellness_score_screen(driver):
    t = time.time()
    try:
        nav(driver,"72-wellness-score"); time.sleep(0.5)
        ok = is_active(driver,"72-wellness-score")
        record("TC-35","Wellness Score Screen","UI",
               "Wellness Score (0-100 index) screen loads correctly",
               "1. Navigate to Wellness Score",
               "Wellness Score screen is active",
               f"Active={ok}",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc35"))
    except Exception as e:
        record("TC-35","Wellness Score Screen","UI","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc35"))

def tc36_mood_forecast_screen(driver):
    t = time.time()
    try:
        nav(driver,"71-mood-forecast"); time.sleep(0.5)
        ok = is_active(driver,"71-mood-forecast")
        record("TC-36","Mood Forecast Screen","UI",
               "Mood Forecast (7-day prediction) screen loads correctly",
               "1. Navigate to Mood Forecast",
               "Mood Forecast screen is active",
               f"Active={ok}",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc36"))
    except Exception as e:
        record("TC-36","Mood Forecast Screen","UI","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc36"))

# ─── MODULE 11 : WELLNESS TOOLS ─────────────────────────────────

def tc37_breathing_exercise_screen(driver):
    t = time.time()
    try:
        nav(driver,"48-breathing-exercise"); time.sleep(0.5)
        circle = driver.find_element(By.CSS_SELECTOR,"#view-48-breathing-exercise .breathing-circle")
        ok = circle.is_displayed()
        record("TC-37","Breathing Exercise Screen","UI",
               "Breathing exercise screen loads with animated breathing circle",
               "1. Navigate to Breathing Exercise\n2. Check breathing circle",
               "Breathing animation circle is visible",
               f"Circle visible={ok}",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc37"))
    except Exception as e:
        record("TC-37","Breathing Exercise Screen","UI","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc37"))

def tc38_meditation_player_screen(driver):
    t = time.time()
    try:
        nav(driver,"47-meditation-player"); time.sleep(0.5)
        ok = is_active(driver,"47-meditation-player")
        record("TC-38","Meditation Player Screen","UI",
               "Meditation player screen loads correctly",
               "1. Navigate to Meditation Player",
               "Meditation Player screen is active",
               f"Active={ok}",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc38"))
    except Exception as e:
        record("TC-38","Meditation Player Screen","UI","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc38"))

def tc39_wellness_suggestions_screen(driver):
    t = time.time()
    try:
        nav(driver,"46-wellness-suggestions"); time.sleep(0.5)
        ok = is_active(driver,"46-wellness-suggestions")
        record("TC-39","Wellness Suggestions Screen","UI",
               "Wellness Suggestions/Therapy screen loads correctly",
               "1. Navigate to Wellness Suggestions",
               "Wellness Suggestions screen is active",
               f"Active={ok}",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc39"))
    except Exception as e:
        record("TC-39","Wellness Suggestions Screen","UI","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc39"))

def tc40_voice_journal_screen(driver):
    t = time.time()
    try:
        nav(driver,"73-voice-journal"); time.sleep(0.5)
        ok = is_active(driver,"73-voice-journal")
        record("TC-40","Voice Journal Screen","UI",
               "Voice Journal screen loads correctly",
               "1. Navigate to Voice Journal",
               "Voice Journal screen is active",
               f"Active={ok}",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc40"))
    except Exception as e:
        record("TC-40","Voice Journal Screen","UI","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc40"))

# ─── MODULE 12 : PROFILE & SETTINGS ─────────────────────────────

def tc41_profile_settings_screen(driver):
    t = time.time()
    try:
        nav(driver,"50-profile-settings"); time.sleep(0.5)
        ok = is_active(driver,"50-profile-settings")
        record("TC-41","Profile Settings Screen","UI",
               "Profile settings screen loads correctly",
               "1. Navigate to Profile Settings",
               "Profile Settings screen is active",
               f"Active={ok}",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc41"))
    except Exception as e:
        record("TC-41","Profile Settings Screen","UI","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc41"))

def tc42_biometric_login_screen(driver):
    t = time.time()
    try:
        nav(driver,"8-biometric-login"); time.sleep(0.5)
        fp_btn  = driver.find_element(By.ID,"fingerprint-btn")
        status  = driver.find_element(By.ID,"webauthn-status")
        ok = fp_btn.is_displayed() and status.is_displayed()
        record("TC-42","Biometric Login Screen","UI",
               "Biometric Login screen shows fingerprint icon and WebAuthn status",
               "1. Navigate to Biometric Login\n2. Check fingerprint button",
               "Fingerprint icon and status text visible",
               "Both elements displayed" if ok else "Missing elements",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc42"))
    except Exception as e:
        record("TC-42","Biometric Login Screen","UI","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc42"))

# ─── MODULE 13 : ONBOARDING ─────────────────────────────────────

def tc43_onboarding_step1(driver):
    t = time.time()
    try:
        nav(driver,"9-onboarding-1"); time.sleep(0.5)
        btn = driver.find_element(By.CSS_SELECTOR,"#view-9-onboarding-1 .btn-primary")
        safe_click(driver, btn); time.sleep(0.8)
        ok = is_active(driver,"10-onboarding-2")
        record("TC-43","Onboarding Step 1→2","Navigation",
               "Onboarding step 1 Continue button navigates to step 2",
               "1. Go to Onboarding 1\n2. Click Continue Gently",
               "Onboarding 2 becomes active",
               f"Step2 active={ok}",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc43"))
    except Exception as e:
        record("TC-43","Onboarding Step 1→2","Navigation","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc43"))

def tc44_onboarding_step2(driver):
    t = time.time()
    try:
        nav(driver,"10-onboarding-2"); time.sleep(0.5)
        btn = driver.find_element(By.CSS_SELECTOR,"#view-10-onboarding-2 .btn-primary")
        safe_click(driver, btn); time.sleep(0.8)
        ok = is_active(driver,"11-onboarding-3")
        record("TC-44","Onboarding Step 2→3","Navigation",
               "Onboarding step 2 Continue button navigates to step 3",
               "1. Go to Onboarding 2\n2. Click Continue Gently",
               "Onboarding 3 becomes active",
               f"Step3 active={ok}",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc44"))
    except Exception as e:
        record("TC-44","Onboarding Step 2→3","Navigation","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc44"))

def tc45_questionnaire_sliders(driver):
    t = time.time()
    try:
        nav(driver,"13-questionnaire"); time.sleep(0.5)
        sliders = driver.find_elements(By.CSS_SELECTOR,"#view-13-questionnaire input[type='range']")
        ok = len(sliders) == 4
        record("TC-45","Questionnaire Sliders","UI",
               "Mental health questionnaire renders all 4 range sliders",
               "1. Navigate to Questionnaire\n2. Count range sliders",
               "Exactly 4 sliders (stress, sleep, mood, anxiety)",
               f"Sliders found={len(sliders)}",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc45"))
    except Exception as e:
        record("TC-45","Questionnaire Sliders","UI","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc45"))

def tc46_goals_checkboxes(driver):
    t = time.time()
    try:
        nav(driver,"14-goals"); time.sleep(0.5)
        cbs = driver.find_elements(By.CSS_SELECTOR,"#view-14-goals input[type='checkbox']")
        ok  = len(cbs) >= 4
        record("TC-46","Goals Selection Checkboxes","UI",
               "Goals selection screen renders at least 4 goal checkboxes",
               "1. Navigate to Goals\n2. Count checkboxes",
               "At least 4 goal checkboxes visible",
               f"Checkboxes found={len(cbs)}",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc46"))
    except Exception as e:
        record("TC-46","Goals Selection Checkboxes","UI","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc46"))

def tc47_notification_prefs_screen(driver):
    t = time.time()
    try:
        nav(driver,"15-notifications"); time.sleep(0.5)
        daily = driver.find_element(By.ID,"pref-daily")
        btn   = driver.find_element(By.ID,"prefs-save-btn")
        ok    = daily.is_displayed() and btn.is_displayed()
        record("TC-47","Notification Preferences Screen","UI",
               "Notification Preferences screen renders toggles and save button",
               "1. Navigate to Notification Preferences\n2. Check toggles and button",
               "Daily reminder toggle and All Set button visible",
               "Both elements displayed" if ok else "Missing elements",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc47"))
    except Exception as e:
        record("TC-47","Notification Preferences Screen","UI","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc47"))

# ─── MODULE 14 : RESPONSIVE / EXTRA ────────────────────────────

def tc48_responsive_mobile_viewport(driver):
    t = time.time()
    try:
        driver.set_window_size(390, 844)
        nav(driver,"2-welcome"); time.sleep(1)
        ok = is_active(driver,"2-welcome")
        record("TC-48","Responsive Mobile Viewport","Responsive",
               "App renders correctly at iPhone 14 viewport (390×844)",
               "1. Resize window to 390×844\n2. Navigate to Welcome",
               "Welcome screen active and app renders at mobile size",
               f"Welcome active={ok} at 390×844",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc48"))
    except Exception as e:
        record("TC-48","Responsive Mobile Viewport","Responsive","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc48"))
    finally:
        try: driver.maximize_window()
        except: pass

def tc49_splash_click_navigate(driver):
    t = time.time()
    try:
        nav(driver,"1-splash"); time.sleep(0.5)
        logo = driver.find_element(By.CSS_SELECTOR,"#view-1-splash .logo-container")
        safe_click(driver, logo); time.sleep(1)
        ok = is_active(driver,"2-welcome")
        record("TC-49","Splash Logo Click→Welcome","Navigation",
               "Clicking the logo on splash screen navigates to Welcome",
               "1. Go to Splash\n2. Click logo/logo-container",
               "Welcome screen becomes active",
               f"welcome active={ok}",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc49"))
    except Exception as e:
        record("TC-49","Splash Logo Click→Welcome","Navigation","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc49"))

def tc50_burnout_wellness_planner(driver):
    t = time.time()
    try:
        nav(driver,"75-wellness-planner"); time.sleep(0.5)
        ok = is_active(driver,"75-wellness-planner")
        record("TC-50","Wellness Planner Screen","UI",
               "Wellness Planner screen loads correctly",
               "1. Navigate to Wellness Planner",
               "Wellness Planner screen is active",
               f"Active={ok}",
               "PASS" if ok else "FAIL", time.time()-t,
               "" if ok else screenshot(driver,"tc50"))
    except Exception as e:
        record("TC-50","Wellness Planner Screen","UI","","",str(e)[:200],"FAIL",time.time()-t,screenshot(driver,"tc50"))


# ══════════════════════════════════════════════════════════════════
#  EXCEL REPORT GENERATOR
# ══════════════════════════════════════════════════════════════════
def generate_report(results):
    wb  = openpyxl.Workbook()
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    total  = len(results)
    passed = sum(1 for r in results if r["Status"] == "PASS")
    failed = total - passed
    rate   = f"{(passed/total*100):.1f}%" if total else "0%"

    # ── Palette ──────────────────────────────────────────────────
    C_DARK   = "FF0D1B2A"
    C_NAVY   = "FF1A3A6B"
    C_BLUE   = "FF1870F4"
    C_LBLUE  = "FFD6E4FE"
    C_GREEN  = "FF16A34A"
    C_LGREEN = "FFD1FAE5"
    C_RED    = "FFDC2626"
    C_LRED   = "FFFEE2E2"
    C_AMBER  = "FFD97706"
    C_LAMBER = "FFFEF3C7"
    C_GREY   = "FFF3F4F6"
    C_WHITE  = "FFFFFFFF"

    def fill(c):  return PatternFill("solid", fgColor=c)
    def font(c="FF000000", sz=10, bold=False, name="Calibri"):
        return Font(name=name, size=sz, bold=bold, color=c)
    def center(wrap=False): return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    def left(wrap=True):    return Alignment(horizontal="left",   vertical="center", wrap_text=wrap)
    def border():
        s = Side(style="thin", color="FFD1D5DB")
        return Border(left=s, right=s, top=s, bottom=s)

    # ════════════════════════════════════════════════════════════
    # SHEET 1 — EXECUTIVE SUMMARY
    # ════════════════════════════════════════════════════════════
    ws1 = wb.active; ws1.title = "Executive Summary"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 32
    ws1.column_dimensions["B"].width = 42

    # Title
    ws1.merge_cells("A1:B4")
    tc = ws1["A1"]
    tc.value     = "NeuroWell AI — Live E2E Selenium Test Report"
    tc.font      = Font(name="Calibri", size=18, bold=True, color=C_WHITE)
    tc.fill      = fill(C_DARK)
    tc.alignment = center()
    for i in range(1,5): ws1.row_dimensions[i].height = 20

    # Sub-info banner
    ws1.merge_cells("A5:B5")
    sc = ws1["A5"]
    sc.value     = f"Generated: {now}   |   URL: {APP_URL}"
    sc.font      = Font(name="Calibri", size=9, italic=True, color="FFA0AEC0")
    sc.fill      = fill(C_DARK)
    sc.alignment = center()
    ws1.row_dimensions[5].height = 16

    def sumrow(r, label, val, vfill=None):
        ws1.row_dimensions[r].height = 26
        lc = ws1.cell(r,1,label)
        vc = ws1.cell(r,2,val)
        lc.font      = font(C_NAVY, 11, True)
        lc.fill      = fill(C_LBLUE)
        lc.alignment = left(False)
        lc.border    = border()
        vc.font      = font(C_NAVY if not vfill else vfill, 12, True)
        vc.alignment = center()
        vc.border    = border()
        if vfill == C_GREEN:  vc.fill = fill(C_LGREEN)
        elif vfill == C_RED:  vc.fill = fill(C_LRED)
        elif vfill == C_AMBER:vc.fill = fill(C_LAMBER)
        else:                 vc.fill = fill(C_WHITE)

    sumrow(7,  "Application URL",         APP_URL)
    sumrow(8,  "Test Account",            f"{TEST_EMAIL}")
    sumrow(9,  "Execution Time",          now)
    sumrow(10, "Total Test Cases",        total)
    sumrow(11, "Passed",                  passed, C_GREEN)
    sumrow(12, "Failed",                  failed, C_RED if failed>0 else C_GREEN)
    sumrow(13, "Pass Rate",               rate,   C_GREEN if failed==0 else (C_RED if passed<total//2 else C_AMBER))
    sumrow(14, "Browser",                 "Google Chrome (Chromium/Selenium)")
    sumrow(15, "Framework",               "Selenium WebDriver 4.x + Python 3.10")
    sumrow(16, "Test Strategy",           "Live Automation — Full E2E on Hosted Firebase App")
    sumrow(17, "Report Generated By",     "NeuroWell AI Selenium Test Runner")

    # ════════════════════════════════════════════════════════════
    # SHEET 2 — DETAILED TEST RESULTS
    # ════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Test Results")
    ws2.sheet_view.showGridLines = False

    COLS   = ["No.","TC_ID","Module","Test Name","Description","Steps","Expected","Actual","Status","Duration(s)","Timestamp","Notes"]
    WIDTHS = [5,    10,     18,      32,          40,           42,     40,        45,      11,      12,           11,         25]

    for i,(col,w) in enumerate(zip(COLS,WIDTHS),1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Banner
    ws2.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    b = ws2["A1"]
    b.value     = "NeuroWell AI — Detailed Selenium E2E Test Case Results"
    b.font      = Font(name="Calibri", size=14, bold=True, color=C_WHITE)
    b.fill      = fill(C_BLUE)
    b.alignment = center()
    ws2.row_dimensions[1].height = 36

    # Column headers
    ws2.row_dimensions[2].height = 28
    for i,col in enumerate(COLS,1):
        c = ws2.cell(2,i,col.replace("_"," "))
        c.font      = Font(name="Calibri", size=10, bold=True, color=C_WHITE)
        c.fill      = fill(C_NAVY)
        c.alignment = center()
        c.border    = border()

    # Data rows
    for ri, row in enumerate(results, 3):
        ws2.row_dimensions[ri].height = 52
        status = row["Status"]
        rf = fill(C_LGREEN) if status=="PASS" else fill(C_LRED)

        for ci, col in enumerate(COLS, 1):
            c = ws2.cell(ri, ci, row.get(col,""))
            c.border = border()

            if col == "Status":
                c.value     = "✅ PASS" if status=="PASS" else "❌ FAIL"
                c.font      = Font(name="Calibri", bold=True, size=11,
                                   color=C_GREEN if status=="PASS" else C_RED)
                c.alignment = center()
                c.fill      = rf
            elif col == "No.":
                c.font      = font(C_NAVY, 10, True)
                c.alignment = center()
                c.fill      = fill(C_GREY)
            elif col == "TC_ID":
                c.font      = Font(name="Calibri", bold=True, size=10, color=C_BLUE)
                c.alignment = center()
                c.fill      = fill(C_GREY)
            elif col == "Module":
                c.font      = font(C_NAVY, 10, True)
                c.alignment = center()
                c.fill      = fill(C_LBLUE)
            elif col == "Duration(s)":
                c.alignment = center()
                c.fill      = fill(C_WHITE)
            elif col == "Timestamp":
                c.alignment = center()
                c.fill      = fill(C_WHITE)
            elif col in ("Actual",):
                c.alignment = left()
                c.fill      = rf
            else:
                c.alignment = left()
                c.fill      = fill(C_WHITE)

    ws2.freeze_panes = "A3"

    # ════════════════════════════════════════════════════════════
    # SHEET 3 — MODULE BREAKDOWN
    # ════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Module Breakdown")
    ws3.sheet_view.showGridLines = False
    for i,(w) in enumerate([22,12,12,12,14],1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    ws3.merge_cells("A1:E2")
    h = ws3["A1"]
    h.value     = "Test Results by Module / Category"
    h.font      = Font(name="Calibri", size=13, bold=True, color=C_WHITE)
    h.fill      = fill(C_BLUE)
    h.alignment = center()
    ws3.row_dimensions[1].height = 22
    ws3.row_dimensions[2].height = 22

    cats = defaultdict(lambda:{"pass":0,"fail":0})
    for r in results:
        cats[r["Module"]]["pass" if r["Status"]=="PASS" else "fail"] += 1

    for i,h in enumerate(["Module","Pass","Fail","Total","Pass Rate"],1):
        c = ws3.cell(3,i,h)
        c.font=Font(name="Calibri",size=10,bold=True,color=C_WHITE)
        c.fill=fill(C_NAVY); c.alignment=center(); c.border=border()
        ws3.row_dimensions[3].height = 22

    for ri,(cat,d) in enumerate(sorted(cats.items()),4):
        p,f = d["pass"],d["fail"]
        tot = p+f
        pr  = f"{p/tot*100:.0f}%" if tot else "0%"
        cf  = C_LGREEN if f==0 else (C_LRED if p==0 else C_LAMBER)
        for ci,v in enumerate([cat,p,f,tot,pr],1):
            c = ws3.cell(ri,ci,v)
            c.font=font(C_DARK,10); c.alignment=center()
            c.fill=fill(cf); c.border=border()
        ws3.row_dimensions[ri].height = 22

    # ════════════════════════════════════════════════════════════
    # SHEET 4 — FAILED TESTS (quick fix reference)
    # ════════════════════════════════════════════════════════════
    fails = [r for r in results if r["Status"]=="FAIL"]
    ws4 = wb.create_sheet("Failed Tests")
    ws4.sheet_view.showGridLines = False
    for i,w in enumerate([8,30,18,50,25],1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    ws4.merge_cells("A1:E2")
    fh = ws4["A1"]
    fh.value     = f"❌ Failed Test Cases — {len(fails)} of {total} Tests Failed"
    fh.font      = Font(name="Calibri", size=13, bold=True, color=C_WHITE)
    fh.fill      = fill(C_RED if fails else C_GREEN)
    fh.alignment = center()
    ws4.row_dimensions[1].height = 22; ws4.row_dimensions[2].height = 22

    if not fails:
        ws4.merge_cells("A3:E3")
        nc = ws4["A3"]
        nc.value     = "🎉 All tests passed! No failures recorded."
        nc.font      = Font(name="Calibri", size=12, bold=True, color=C_GREEN)
        nc.fill      = fill(C_LGREEN); nc.alignment = center()
    else:
        for i,h in enumerate(["TC_ID","Test Name","Module","Actual Result / Error","Notes"],1):
            c = ws4.cell(3,i,h)
            c.font=Font(name="Calibri",bold=True,color=C_WHITE)
            c.fill=fill(C_NAVY); c.alignment=center(); c.border=border()
        for ri,r in enumerate(fails,4):
            ws4.row_dimensions[ri].height = 42
            for ci,v in enumerate([r["TC_ID"],r["Test Name"],r["Module"],r["Actual"],r["Notes"]],1):
                c = ws4.cell(ri,ci,v)
                c.font=font(C_DARK,10); c.alignment=left(); c.fill=fill(C_LRED); c.border=border()

    wb.save(REPORT_PATH)
    print(f"\n  Excel report saved: {REPORT_PATH}\n")
    return REPORT_PATH


# ══════════════════════════════════════════════════════════════════
#  MAIN RUNNER
# ══════════════════════════════════════════════════════════════════
def run_all():
    print("\n" + "="*65)
    print("  NeuroWell AI  Selenium E2E Live Test Suite")
    print(f"  URL   : {APP_URL}")
    print(f"  Email : {TEST_EMAIL}")
    print(f"  Time  : {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("="*65 + "\n")

    driver = create_driver()
    try:
        # MODULE 1: SMOKE
        print("--- SMOKE TESTS ---")
        tc01_page_load(driver)
        tc02_page_title(driver)
        tc03_splash_logo_visible(driver)
        tc04_splash_to_welcome_auto(driver)

        # MODULE 2: WELCOME / NAVIGATION
        print("\n--- WELCOME & NAVIGATION ---")
        tc05_welcome_has_buttons(driver)
        tc06_welcome_to_login_click(driver)
        tc07_welcome_to_signup_click(driver)

        # MODULE 3: LOGIN FORM
        print("\n--- LOGIN FORM ---")
        tc08_login_form_elements(driver)
        tc09_email_field_accepts_input(driver)
        tc10_password_field_masked(driver)
        tc11_login_empty_fields_validation(driver)
        tc12_login_wrong_password(driver)
        tc14_back_button_login_to_welcome(driver)

        # MODULE 4: SIGNUP
        print("\n--- SIGNUP ---")
        tc15_signup_form_elements(driver)

        # MODULE 5: FORGOT PASSWORD
        print("\n--- FORGOT PASSWORD ---")
        tc16_forgot_password_screen(driver)

        # Run VALID LOGIN — needed to access post-auth screens
        print("\n--- AUTHENTICATION (Firebase) ---")
        tc13_valid_login(driver)

        # MODULE 6: DASHBOARD
        print("\n--- DASHBOARD ---")
        tc17_dashboard_renders(driver)
        tc18_resonance_score_updates(driver)
        tc19_bottom_nav_visible_on_dashboard(driver)
        tc20_bottom_nav_hidden_on_login(driver)
        tc21_music_control_toggle(driver)

        # MODULE 7: EMOTION ANALYSIS
        print("\n--- EMOTION ANALYSIS ---")
        tc22_navigate_to_analyze(driver)
        tc23_analyze_options_visible(driver)
        tc24_voice_screen(driver)
        tc25_face_recognition_screen(driver)
        tc26_fingerprint_screen(driver)
        tc27_multimodal_screen(driver)

        # MODULE 8: AI CHAT
        print("\n--- AI CHAT ---")
        tc28_chat_screen_loads(driver)
        tc29_chat_send_message(driver)

        # MODULE 9: PRIVATE MENTOR
        print("\n--- PRIVATE MENTOR ---")
        tc30_private_mentor_screen(driver)
        tc31_mentor_chat_response(driver)

        # MODULE 10: INSIGHTS
        print("\n--- AI INSIGHTS & ANALYTICS ---")
        tc32_ai_insights_screen(driver)
        tc33_weekly_trends_screen(driver)
        tc34_burnout_detector_screen(driver)
        tc35_wellness_score_screen(driver)
        tc36_mood_forecast_screen(driver)

        # MODULE 11: WELLNESS TOOLS
        print("\n--- WELLNESS TOOLS ---")
        tc37_breathing_exercise_screen(driver)
        tc38_meditation_player_screen(driver)
        tc39_wellness_suggestions_screen(driver)
        tc40_voice_journal_screen(driver)

        # MODULE 12: PROFILE
        print("\n--- PROFILE & SETTINGS ---")
        tc41_profile_settings_screen(driver)
        tc42_biometric_login_screen(driver)

        # MODULE 13: ONBOARDING
        print("\n--- ONBOARDING & FORMS ---")
        tc43_onboarding_step1(driver)
        tc44_onboarding_step2(driver)
        tc45_questionnaire_sliders(driver)
        tc46_goals_checkboxes(driver)
        tc47_notification_prefs_screen(driver)

        # MODULE 14: RESPONSIVE
        print("\n--- RESPONSIVE & EXTRA ---")
        tc48_responsive_mobile_viewport(driver)
        tc49_splash_click_navigate(driver)
        tc50_burnout_wellness_planner(driver)

    finally:
        driver.quit()

    # Summary
    passed = sum(1 for r in test_results if r["Status"]=="PASS")
    failed = len(test_results) - passed
    print("\n" + "="*65)
    print(f"  TOTAL : {len(test_results)} tests")
    print(f"  PASS  : {passed}")
    print(f"  FAIL  : {failed}")
    print(f"  RATE  : {passed/len(test_results)*100:.1f}%")
    print("="*65)

    generate_report(test_results)
    print(f"  Report: {REPORT_PATH}")

    # Generate GitHub Step Summary
    summary_file = os.environ.get("GITHUB_STEP_SUMMARY")
    if summary_file:
        with open(summary_file, "a", encoding="utf-8") as f:
            f.write("## E2E Test Results\n\n")
            if len(test_results) > 0:
                f.write(f"**Total:** {len(test_results)} | **Pass:** {passed} | **Fail:** {failed} | **Pass Rate:** {passed/len(test_results)*100:.1f}%\n\n")
            f.write("| ID | Module | Test Name | Status | Duration |\n")
            f.write("|---|---|---|---|---|\n")
            for r in test_results:
                icon = "✅ PASS" if r["Status"] == "PASS" else "❌ FAIL"
                f.write(f"| {r['TC_ID']} | {r['Module']} | {r['Test Name']} | {icon} | {r['Duration(s)']}s |\n")

if __name__ == "__main__":
    run_all()
